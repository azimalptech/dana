"""Applies an edited Interface-Texts.xlsx back into lib/core/l10n.dart.

Surgical, not regenerating: for each key it rewrites only the tk/ru/en
string literals in place, so the comments that document where each string
appears on screen survive untouched.

    python tool/apply_strings.py ../Interface-Texts.xlsx          # dry run
    python tool/apply_strings.py ../Interface-Texts.xlsx --write

Refuses to write when the sheet breaks a format rule — a changed number of
`|` forms or a dropped `{placeholder}` ships a visibly broken string
("{n} of {m}" losing {m}), and those are exactly the edits a reviewer makes
by accident.
"""
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

L10N = Path(__file__).resolve().parent.parent / 'lib' / 'core' / 'l10n.dart'
LANGS = ('tk', 'ru', 'en')
PLACEHOLDER = re.compile(r'\{[a-z]\}')


def dart_literal(value: str) -> str:
    """A single-quoted Dart literal. `$` starts interpolation, so it escapes."""
    body = value.replace('\\', '\\\\').replace("'", "\\'").replace('$', r'\$')
    return f"'{body}'"


def decode_literal(literal: str) -> str:
    """The value a Dart literal denotes.

    Long strings are written as adjacent literals across lines, which Dart
    concatenates — so the source text differs from a single-literal
    rewrite while the value is identical. Comparing decoded values keeps an
    untouched row from being reported as an edit.
    """
    parts = re.findall(r"'((?:[^'\\]|\\.)*)'", literal)
    joined = ''.join(parts)
    return (joined.replace("\\'", "'")
                  .replace(r'\$', '$')
                  .replace('\\n', '\n')
                  .replace('\\\\', '\\'))


def read_sheet(path: Path) -> dict:
    """Reads the review sheet, locating columns by their header.

    The sheet has been through more than one layout — key-first, and later
    grouped by screen with Screen/Where columns in front — so the columns
    are found by name rather than position. Screen-band rows carry no key
    and are skipped.
    """
    ws = load_workbook(path, data_only=True)['Interface texts']

    header = {}
    for cell in ws[1]:
        if cell.value:
            header[str(cell.value).strip().lower()] = cell.column - 1

    def col(*names):
        for n in names:
            for text, idx in header.items():
                if text.startswith(n):
                    return idx
        raise SystemExit(f'column {names[0]!r} not found — is this the right sheet?')

    c_key, c_en, c_tk, c_ru = (col('key'), col('english'), col('turkmen'), col('russian'))

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[c_key] if c_key < len(row) else None
        if key is None or not str(key).strip():
            continue                      # screen band, or a blank spacer
        out[str(key).strip()] = {
            'en': '' if row[c_en] is None else str(row[c_en]),
            'tk': '' if row[c_tk] is None else str(row[c_tk]),
            'ru': '' if row[c_ru] is None else str(row[c_ru]),
        }
    return out


def current_values(source: str) -> dict:
    """Every key's existing literals, read straight out of the Dart map."""
    out = {}
    for m in re.finditer(r"^    '([a-z0-9_]+)':\s*\{", source, re.M):
        key = m.group(1)
        depth, i = 0, m.end() - 1
        while i < len(source):
            if source[i] == '{':
                depth += 1
            elif source[i] == '}':
                depth -= 1
                if depth == 0:
                    break
            i += 1
        out[key] = (m.end() - 1, i + 1, source[m.end() - 1:i + 1])
    return out


def validate(sheet: dict, existing: dict) -> list:
    problems = []

    unknown = set(sheet) - set(existing)
    missing = set(existing) - set(sheet)
    for k in sorted(unknown):
        problems.append(f'{k}: not a key in l10n.dart (was the Key column edited?)')
    for k in sorted(missing):
        problems.append(f'{k}: row missing from the sheet (rows must not be deleted)')

    for key, vals in sheet.items():
        if key not in existing:
            continue
        block = existing[key][2]
        for lang in LANGS:
            new = vals[lang]
            if not new.strip():
                problems.append(f'{key}.{lang}: empty')
                continue
            old_m = re.search(rf"'{lang}':\s*((?:'(?:[^'\\]|\\.)*'\s*)+)", block)
            if old_m is None:
                continue
            old = decode_literal(old_m.group(1))

            if old.count('|') != new.count('|'):
                problems.append(
                    f'{key}.{lang}: has {new.count("|")} "|" but needs '
                    f'{old.count("|")} (forms must stay the same count)')

            want = set(PLACEHOLDER.findall(old))
            got = set(PLACEHOLDER.findall(new))
            if want - got:
                problems.append(
                    f'{key}.{lang}: lost placeholder(s) {sorted(want - got)}')
            if got - want:
                problems.append(
                    f'{key}.{lang}: unknown placeholder(s) {sorted(got - want)}')
    return problems


def apply(source: str, sheet: dict, existing: dict) -> tuple:
    changed = []
    # Rewrite from the bottom so earlier offsets stay valid.
    for key in sorted(existing, key=lambda k: existing[k][0], reverse=True):
        if key not in sheet:
            continue
        start, end, block = existing[key]
        new_block = block
        for lang in LANGS:
            m = re.search(rf"('{lang}':\s*)((?:'(?:[^'\\]|\\.)*'\s*)+)", new_block)
            if m is None:
                continue
            old_literal = m.group(2).rstrip()
            new_literal = dart_literal(sheet[key][lang])
            # Compare values, not source text: an untouched multi-line
            # literal must not read as an edit.
            if decode_literal(old_literal) == sheet[key][lang]:
                continue
            new_block = new_block[:m.start(2)] + new_literal + new_block[m.end(2):]
            changed.append(f'{key}.{lang}')
        if new_block != block:
            source = source[:start] + new_block + source[end:]
    return source, changed


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2

    xlsx = Path(sys.argv[1])
    write = '--write' in sys.argv

    # newline='' keeps the file's own line endings out of Python's
    # translation, so applying two edits does not rewrite all 758 lines.
    with open(L10N, encoding='utf-8', newline='') as fh:
        source = fh.read()
    sheet = read_sheet(xlsx)
    existing = current_values(source)

    print(f'sheet rows: {len(sheet)}   dart keys: {len(existing)}')

    problems = validate(sheet, existing)
    if problems:
        print(f'\n{len(problems)} problem(s) — nothing written:\n')
        for p in problems:
            print('  •', p)
        return 1

    updated, changed = apply(source, sheet, existing)

    print(f'changed: {len(changed)} value(s)')
    for c in changed[:40]:
        print('  -', c)
    if len(changed) > 40:
        print(f'  … and {len(changed) - 40} more')

    if not changed:
        print('nothing to do')
        return 0

    if write:
        with open(L10N, 'w', encoding='utf-8', newline='') as fh:
            fh.write(updated)
        print(f'\nwritten -> {L10N}')
        print('now run: flutter analyze lib/core/l10n.dart')
    else:
        print('\ndry run — pass --write to apply')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
